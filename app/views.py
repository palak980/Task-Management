from django.shortcuts import render
from django.core.paginator import Paginator
from django.http import HttpResponse
from .models import User, Task
from .forms import UserForm, TaskForm
# Create your views here.
def add_user(request):
    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            form.save()
            return HttpResponse('User added successfully!')
    else:
        form = UserForm()
    return render(request, 'add_user.html', {'form': form})

def add_task(request):
    if request.method == 'POST':
        form = TaskForm(request.POST)
        if form.is_valid():
            form.save()
            return HttpResponse('Task added successfully!')
    else:
        form = TaskForm()
    return render(request, 'add_task.html', {'form': form})

def user_list(request):
    users = User.objects.all()
    paginator = Paginator(users, 10) # Show 10 users per page.
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'user_list.html', {'page_obj': page_obj})

def task_list(request):
    tasks = Task.objects.all()
    paginator = Paginator(tasks, 10) # Show 10 tasks per page.
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'task_list.html', {'page_obj': page_obj})

def export_to_excel(request):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.writer.excel import save_virtual_workbook

    workbook = openpyxl.Workbook()
    user_sheet = workbook.active
    user_sheet.title = 'Users'
    task_sheet = workbook.create_sheet(title='Tasks')

    # Writing users data to excel sheet
    users = User.objects.all()
    user_sheet.append(['ID', 'Name', 'Email', 'Mobile'])
    for user in users:
        user_sheet.append([user.user_id, user.name, user.email, user.mobile])

    # Writing tasks data to excel sheet
    tasks = Task.objects.all()
    task_sheet.append(['User ID', 'Task Detail', 'Task Type'])
    for task in tasks:
        task_sheet.append([task.user.user_id, task.task_detail, task.task_type])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=users_tasks.xlsx'
    response.write(save_virtual_workbook(workbook))
    return response
